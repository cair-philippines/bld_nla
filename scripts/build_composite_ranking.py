"""
Build the composite priority ranking Excel file.

Ranks schools based on the composite of both within-school-year Learning
segments (SY 2024-25 and SY 2025-26). Only schools with valid_strict data
in both segments are ranked. First-cycle schools (131) are tagged in the
full ranking and excluded from the Top 100.

Usage:
    cd /workspace/innovation-projects/project_crla
    ds python scripts/build_composite_ranking.py

Output:
    output/priority_ranking_composite.xlsx
      - Sheet 1: All Schools (Ranked)
      - Sheet 2: Top 100 (Priority Schools)
      - Sheet 3: All Schools (Reference)
      - Sheet 4: Notes
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment

# Add modules to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "modules"))

from preprocessing import (
    resolve_latest_exports, load_all_assessments, get_total_assessed,
    CANONICAL_GRADE_COLUMNS, _clean_raw_to_numeric,
)
from analysis import (
    process_all_timepoints, compute_chain_progress,
    _build_segment_pairs, _segment_label,
)
from lgu_matching import (
    load_deped_psgc, build_school_lgu_crosswalk,
    load_lgu_revenue, match_lgu_revenue,
)
from priority_ranking import _z_score, DEFAULT_NEED_WEIGHTS

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

FIRST_CYCLE_PATH = PROJECT_ROOT / "output" / "bbi_annex_b_school_ids.txt"
OUTPUT_PATH = PROJECT_ROOT / "output" / "priority_ranking_composite.xlsx"

PSGC_PATH = PROJECT_ROOT / "data/raw/SY 2024-2025 School Level Database WITH PSGC.xlsx"
BLGF_PATH = PROJECT_ROOT / "data/raw/By-LGU-SRE-2024.xlsx"
ENROLLMENT_PATH = PROJECT_ROOT / "data/raw/public_project_bukas_enrollment_2024-25.csv"


_TRANS_COLS = [c for c in CANONICAL_GRADE_COLUMNS if "Transitioning" in c]
_GL_COLS = [c for c in CANONICAL_GRADE_COLUMNS if "Grade Level" in c]


def _count_trans_plus(raw_df):
    """Count students at Transitioning or Grade Level per school (from raw counts)."""
    clean = _clean_raw_to_numeric(raw_df)
    return clean[_TRANS_COLS + _GL_COLS].sum(axis=1)


def _count_gl(raw_df):
    """Count students at Grade Level per school (from raw counts)."""
    clean = _clean_raw_to_numeric(raw_df)
    return clean[_GL_COLS].sum(axis=1)


def _assign_band(p):
    if p >= 0.9:
        return "90th+"
    elif p >= 0.7:
        return "70–90th"
    elif p >= 0.5:
        return "50–70th"
    else:
        return "Below 50th"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # ── Load 131 first-cycle IDs ─────────────────────────────────────────
    with open(FIRST_CYCLE_PATH) as f:
        first_cycle_ids = set(int(line.strip()) for line in f if line.strip())
    print(f"1st cycle schools: {len(first_cycle_ids)}")

    # ── Load and process ─────────────────────────────────────────────────
    file_map = resolve_latest_exports()
    df_all = load_all_assessments(file_map=file_map, source="local")
    results = process_all_timepoints(df_all, scoring="ordinal")
    progress_df = compute_chain_progress(
        performance=results["performance"],
        raw_data=df_all,
        validation=results["validation"],
        ordinal_sd=results["ordinal_sd"],
        ordinal_skew=results["ordinal_skew"],
    )

    # LGU data
    psgc_df = load_deped_psgc(str(PSGC_PATH))
    crosswalk_df = build_school_lgu_crosswalk(psgc_df)
    lgu_revenue_df = load_lgu_revenue(str(BLGF_PATH))
    matched_lgu_df, _, _ = match_lgu_revenue(crosswalk_df, lgu_revenue_df)

    # Enrollment
    enroll = pd.read_csv(str(ENROLLMENT_PATH))
    grade_cols = [c for c in enroll.columns if c.endswith("_male") or c.endswith("_female")]
    enroll["total_enrolled"] = enroll[grade_cols].sum(axis=1)

    # ── Identify segment columns ─────────────────────────────────────────
    pairs = _build_segment_pairs()
    pair_labels = {_segment_label(t0, t1): (i, t0, t1) for i, (t0, t1) in enumerate(pairs)}

    seg1_idx, seg1_t0, seg1_t1 = pair_labels["Learning_2024-25"]
    seg4_idx, seg4_t0, seg4_t1 = pair_labels["Learning_2025-26"]
    seg1_n = seg1_idx + 1
    seg4_n = seg4_idx + 1

    strict_col_1 = f"seg{seg1_n}_valid_strict"
    strict_col_4 = f"seg{seg4_n}_valid_strict"
    delta_col_1 = f"seg{seg1_n}_Learning_2024-25"
    delta_col_4 = f"seg{seg4_n}_Learning_2025-26"
    sd_delta_col_1 = f"seg{seg1_n}_Learning_2024-25_sd_delta"
    sd_delta_col_4 = f"seg{seg4_n}_Learning_2025-26_sd_delta"
    skew_delta_col_1 = f"seg{seg1_n}_Learning_2024-25_skew_delta"
    skew_delta_col_4 = f"seg{seg4_n}_Learning_2025-26_skew_delta"

    # ── Filter to schools valid in BOTH Learning segments ────────────────
    both_strict = progress_df[strict_col_1] & progress_df[strict_col_4]
    eligible = progress_df[both_strict].copy()
    print(f"\nSchools valid in both Learning segments (strict): {len(eligible)}")

    # ── Build ranking DataFrame ──────────────────────────────────────────
    df = pd.DataFrame(index=eligible.index)
    df.index.name = "School ID"

    df["school_name"] = eligible.get("School Name")
    df["division"] = eligible.get("Division")
    df["region"] = eligible.get("Region")

    # Province and Municipality from crosswalk
    xw = crosswalk_df.copy()
    if "School ID" in xw.columns:
        xw = xw.set_index("School ID")
    df["province"] = xw["psgc_province"].reindex(df.index)
    df["municipality"] = xw["psgc_muni_name"].reindex(df.index)

    # Latest endpoint: EoSY 2025-26
    df["mean_end"] = eligible["perf_EoSY_2025-26"]
    df["sd_end"] = eligible["sd_EoSY_2025-26"]
    df["skew_end"] = eligible["skew_EoSY_2025-26"]

    # Per-year deltas
    df["delta_mean_2024_25"] = eligible[delta_col_1]
    df["delta_mean_2025_26"] = eligible[delta_col_4]

    # Average deltas for Need composite
    df["delta_mean_avg"] = (df["delta_mean_2024_25"] + df["delta_mean_2025_26"]) / 2
    df["delta_sd_avg"] = (eligible[sd_delta_col_1].fillna(0) + eligible[sd_delta_col_4].fillna(0)) / 2
    df["delta_skew_avg"] = (eligible[skew_delta_col_1].fillna(0) + eligible[skew_delta_col_4].fillna(0)) / 2

    # Impact: assessed count at latest endpoint (EoSY 2025-26)
    assessed = get_total_assessed("2025-26", "EoSY")
    df["assessed_count"] = assessed.reindex(df.index)

    # Net gain columns: Trans+ and GL (net change BoSY→EoSY per school year)
    for sy, sy_label in [("2024-25", "2024_25"), ("2025-26", "2025_26")]:
        bosy_key = (sy, "BoSY")
        eosy_key = (sy, "EoSY")
        if bosy_key in df_all and eosy_key in df_all:
            tp_bosy = _count_trans_plus(df_all[bosy_key]).reindex(df.index)
            tp_eosy = _count_trans_plus(df_all[eosy_key]).reindex(df.index)
            df[f"net_gain_trans_plus_{sy_label}"] = tp_eosy - tp_bosy

            gl_bosy = _count_gl(df_all[bosy_key]).reindex(df.index)
            gl_eosy = _count_gl(df_all[eosy_key]).reindex(df.index)
            df[f"net_gain_gl_{sy_label}"] = gl_eosy - gl_bosy

    # Capacity Gap: LGU SEF per capita
    df["psgc_muni_code"] = xw["psgc_muni_code"].reindex(df.index)

    lgu = matched_lgu_df[["psgc_muni_code", "lgu_name", "rpt_special_education_fund"]].drop_duplicates(
        subset=["psgc_muni_code"]
    )
    code_to_sef = lgu.set_index("psgc_muni_code")["rpt_special_education_fund"]
    code_to_name = lgu.set_index("psgc_muni_code")["lgu_name"]
    df["lgu_name"] = df["psgc_muni_code"].map(code_to_name)
    df["lgu_sef"] = df["psgc_muni_code"].map(code_to_sef)

    enr = enroll[["school_id", "total_enrolled"]].merge(
        xw[["psgc_muni_code"]].reset_index(),
        left_on="school_id", right_on="School ID", how="inner",
    )
    lgu_enrolled = enr.groupby("psgc_muni_code")["total_enrolled"].sum()
    df["lgu_enrolled"] = df["psgc_muni_code"].map(lgu_enrolled)
    df["sef_per_capita"] = df["lgu_sef"] / df["lgu_enrolled"]

    # Drop incomplete
    required = ["mean_end", "delta_mean_avg", "assessed_count", "sef_per_capita"]
    n_before = len(df)
    df = df.dropna(subset=required)
    print(f"Dropped (missing LGU/count data): {n_before - len(df)}")
    print(f"Ranked: {len(df)}")

    for col in ["sd_end", "skew_end", "delta_sd_avg", "delta_skew_avg"]:
        df[col] = df[col].fillna(0)

    # ── Three-pillar ranking ─────────────────────────────────────────────
    need_weights = DEFAULT_NEED_WEIGHTS.copy()

    need_components = {
        "level_mean": _z_score(5 - df["mean_end"]),
        "delta_mean": _z_score(-df["delta_mean_avg"]),
        "level_sd": _z_score(df["sd_end"]),
        "level_skew": _z_score(df["skew_end"]),
        "delta_sd": _z_score(df["delta_sd_avg"]),
        "delta_skew": _z_score(df["delta_skew_avg"]),
    }
    df["need_score"] = sum(need_weights[k] * v for k, v in need_components.items())
    df["impact_score"] = df["assessed_count"]
    df["capacity_gap_score"] = _z_score(-df["sef_per_capita"])

    df["need_pctile"] = df["need_score"].rank(pct=True)
    df["impact_pctile"] = df["impact_score"].rank(pct=True)
    df["capacity_gap_pctile"] = df["capacity_gap_score"].rank(pct=True)

    df["priority_score"] = df["need_pctile"] * df["impact_pctile"] * df["capacity_gap_pctile"]
    df["priority_pctile"] = df["priority_score"].rank(pct=True)
    df = df.sort_values("priority_score", ascending=False)

    df["priority_band"] = df["priority_pctile"].apply(_assign_band)
    df["is_first_cycle"] = df.index.isin(first_cycle_ids)
    n_fc_ranked = df["is_first_cycle"].sum()
    print(f"\n1st cycle schools in ranking: {n_fc_ranked} of {len(first_cycle_ids)}")

    # ── Build output sheets ──────────────────────────────────────────────

    # Sheet 1: All Schools (data sheet)
    ref = pd.DataFrame(index=progress_df.index)
    ref.index.name = "School ID"
    ref["School Name"] = progress_df.get("School Name")
    ref["Division"] = progress_df.get("Division")
    ref["Region"] = progress_df.get("Region")
    ref["Province"] = xw["psgc_province"].reindex(ref.index)
    ref["Municipality"] = xw["psgc_muni_name"].reindex(ref.index)

    ref["Valid SY 2024-25"] = progress_df[strict_col_1].map({True: "Yes", False: "No"})
    ref["Valid SY 2025-26"] = progress_df[strict_col_4].map({True: "Yes", False: "No"})
    ref["Valid for Composite"] = (progress_df[strict_col_1] & progress_df[strict_col_4]).map({True: "Yes", False: "No"})
    ref["1st Cycle"] = pd.Series(ref.index.isin(first_cycle_ids), index=ref.index).map({True: "Yes", False: ""})

    ref["Mean Ordinal (EoSY 2024-25)"] = progress_df.get("perf_EoSY_2024-25")
    ref["Mean Ordinal (EoSY 2025-26)"] = progress_df.get("perf_EoSY_2025-26")
    ref["Delta Mean (SY 2024-25)"] = progress_df.get(delta_col_1)
    ref["Delta Mean (SY 2025-26)"] = progress_df.get(delta_col_4)

    # Assessed counts and net gain columns for reference sheet
    for sy, sy_label in [("2024-25", "2024-25"), ("2025-26", "2025-26")]:
        bosy_key = (sy, "BoSY")
        eosy_key = (sy, "EoSY")
        if bosy_key in df_all and eosy_key in df_all:
            # Raw assessed counts at both endpoints
            bosy_assessed = get_total_assessed(sy, "BoSY").reindex(ref.index)
            eosy_assessed = get_total_assessed(sy, "EoSY").reindex(ref.index)
            ref[f"Assessed (BoSY {sy_label})"] = bosy_assessed
            ref[f"Assessed (EoSY {sy_label})"] = eosy_assessed

            # Net gain (assessment-level: counts across all language groups per grade)
            tp_bosy = _count_trans_plus(df_all[bosy_key]).reindex(ref.index)
            tp_eosy = _count_trans_plus(df_all[eosy_key]).reindex(ref.index)
            ref[f"Net Gain Trans+ Assessments (SY {sy_label})"] = tp_eosy - tp_bosy

            gl_bosy = _count_gl(df_all[bosy_key]).reindex(ref.index)
            gl_eosy = _count_gl(df_all[eosy_key]).reindex(ref.index)
            ref[f"Net Gain GL Assessments (SY {sy_label})"] = gl_eosy - gl_bosy

    # SEF per capita for reference sheet
    ref["psgc_muni_code"] = xw["psgc_muni_code"].reindex(ref.index)
    lgu = matched_lgu_df[["psgc_muni_code", "lgu_name", "rpt_special_education_fund"]].drop_duplicates(
        subset=["psgc_muni_code"]
    )
    code_to_sef = lgu.set_index("psgc_muni_code")["rpt_special_education_fund"]
    code_to_name = lgu.set_index("psgc_muni_code")["lgu_name"]
    ref["LGU Name"] = ref["psgc_muni_code"].map(code_to_name)
    ref_lgu_sef = ref["psgc_muni_code"].map(code_to_sef)
    enr_xw = enroll[["school_id", "total_enrolled"]].merge(
        xw[["psgc_muni_code"]].reset_index(),
        left_on="school_id", right_on="School ID", how="inner",
    )
    lgu_enrolled = enr_xw.groupby("psgc_muni_code")["total_enrolled"].sum()
    ref_lgu_enrolled = ref["psgc_muni_code"].map(lgu_enrolled)
    ref["SEF per Capita"] = ref_lgu_sef / ref_lgu_enrolled
    ref = ref.drop(columns=["psgc_muni_code"])

    # Detailed exclusion reasons from validation flags
    validation = results["validation"]
    count_stable_col_1 = f"seg{seg1_n}_count_stable"
    count_stable_col_4 = f"seg{seg4_n}_count_stable"

    reasons = []
    for sid in progress_df.index:
        v1 = progress_df.at[sid, strict_col_1]
        v4 = progress_df.at[sid, strict_col_4]
        if v1 and v4:
            reasons.append("")
            continue

        parts = []
        for sy, strict_flag, count_col in [
            ("SY 2024-25", strict_col_1, count_stable_col_1),
            ("SY 2025-26", strict_col_4, count_stable_col_4),
        ]:
            if progress_df.at[sid, strict_flag]:
                continue  # this SY is fine

            sy_key_bosy = (sy.replace("SY ", ""), "BoSY")
            sy_key_eosy = (sy.replace("SY ", ""), "EoSY")
            sub = []

            # Check each endpoint's validation
            for tp_key, tp_label in [(sy_key_bosy, "BoSY"), (sy_key_eosy, "EoSY")]:
                if tp_key not in validation:
                    sub.append(f"no {tp_label} data")
                    continue
                val = validation[tp_key]
                if sid not in val.index:
                    sub.append(f"not in {tp_label} dataset")
                    continue
                if not val.at[sid, "valid"]:
                    sub.append(f"no valid groups at {tp_label}")
                    continue
                if not val.at[sid, "has_all_grades"]:
                    sub.append(f"missing grade level(s) at {tp_label}")
                if val.at[sid, "groups_available"] < 4:
                    sub.append(f"only {int(val.at[sid, 'groups_available'])} groups at {tp_label}")
                mga = val.at[sid, "min_group_assessed"]
                if pd.notna(mga) and mga < 15:
                    sub.append(f"min {int(mga)} assessed/group at {tp_label}")

            # Count stability
            if count_col in progress_df.columns:
                if not progress_df.at[sid, count_col]:
                    if not any("no " in s or "not in" in s for s in sub):
                        sub.append("unstable count (>25% change)")

            if sub:
                parts.append(f"{sy}: {'; '.join(sub)}")
            else:
                parts.append(f"{sy}: failed strict validation")

        reasons.append(" | ".join(parts))
    ref["Exclusion Reason"] = reasons
    ref = ref.reset_index()
    ref = ref.sort_values(["Valid for Composite", "1st Cycle", "School Name"], ascending=[False, False, True])

    n_fc_in_ref = (ref["1st Cycle"] == "Yes").sum()
    n_fc_valid = ((ref["1st Cycle"] == "Yes") & (ref["Valid for Composite"] == "Yes")).sum()
    print(f"Reference sheet: {len(ref)} total, {n_fc_in_ref} tagged 1st Cycle ({n_fc_valid} valid)")

    # ── Write Excel ──────────────────────────────────────────────────────
    with pd.ExcelWriter(str(OUTPUT_PATH), engine="openpyxl") as writer:
        ref.to_excel(writer, sheet_name="All Schools", index=False)

        ws = writer.book.create_sheet("Notes")
        n_valid_composite = (ref["Valid for Composite"] == "Yes").sum()
        n_excluded = len(ref) - n_valid_composite
        notes = [
            ("Field", "Value"),
            ("Generated", pd.Timestamp.now().strftime("%Y-%m-%d")),
            ("Data Source", "CRLA dashboard exports (data/raw/dashboard_export/)"),
            ("School Years", "SY 2024-25 and SY 2025-26 (within-school-year evaluation only)"),
            ("Total Schools in Dataset", f"{len(ref):,}"),
            ("Valid for Composite (both SYs)", f"{n_valid_composite:,}"),
            ("Valid SY 2024-25 only", f"{((ref['Valid SY 2024-25'] == 'Yes') & (ref['Valid SY 2025-26'] == 'No')).sum():,}"),
            ("Valid SY 2025-26 only", f"{((ref['Valid SY 2024-25'] == 'No') & (ref['Valid SY 2025-26'] == 'Yes')).sum():,}"),
            ("Failed both SYs", f"{((ref['Valid SY 2024-25'] == 'No') & (ref['Valid SY 2025-26'] == 'No')).sum():,}"),
            ("", ""),
            ("1ST CYCLE SCHOOLS", ""),
            ("Source", "output/bbi_annex_b_school_ids.txt (131 unique School IDs)"),
            ("Tagged in data sheet", f"{n_fc_in_ref} of 131 — '1st Cycle = Yes', filterable"),
            ("Of which valid for composite", f"{n_fc_valid} of 131"),
            ("", ""),
            ("VALIDITY CRITERIA", ""),
            ("Criterion", "Requirement"),
            ("Grade Coverage", "All three grade levels (G1, G2, G3) must have at least one reporting group"),
            ("Group Breadth", "At least 4 of 6 grade-language groups must be reporting"),
            ("Minimum Sample", "At least 15 assessed learners in every reporting group"),
            ("Count Stability", "Total assessed count must not change by more than 25% between BoSY and EoSY"),
            ("Per-SY Validity", "School must pass ALL above criteria at BOTH endpoints (BoSY and EoSY) of a school year"),
            ("Composite Validity", "School must be valid in BOTH SY 2024-25 AND SY 2025-26"),
            ("", ""),
            ("COLUMN DEFINITIONS", ""),
            ("Column", "Interpretation"),
            ("Valid SY 2024-25 / 2025-26", "Yes if school passes strict validation for that school year's Learning segment."),
            ("Valid for Composite", "Yes only if BOTH school years pass."),
            ("1st Cycle", "Yes if school is in the original 131-school list submitted to stakeholders."),
            ("Mean Ordinal (EoSY YYYY)", "Average proficiency level at end of school year (1=Lower Emergent, 5=Grade Level)."),
            ("Delta Mean (SY YYYY)", "Change in ordinal mean from BoSY to EoSY. Positive = improvement."),
            ("Assessed (BoSY/EoSY YYYY)", "Unique assessed students at that endpoint (from dashboard Total Assessed). Compare BoSY vs EoSY to judge count stability."),
            ("Net Gain Trans+ Assessments", "Net change in Trans+ assessment results (across all language groups) from BoSY to EoSY. Counts assessments, not unique students — one student assessed in multiple languages contributes multiple counts. Use for relative comparison between schools."),
            ("Net Gain GL Assessments", "Net change in Grade Level assessment results (across all language groups) from BoSY to EoSY. Same assessment-level counting as Trans+. Use for relative comparison between schools."),
            ("LGU Name", "Local Government Unit matched via PSGC crosswalk."),
            ("SEF per Capita", "LGU Special Education Fund ÷ total enrolled learners (PHP). Lower = less local funding."),
            ("Province / Municipality", "From the DepEd PSGC database."),
            ("Exclusion Reason", "Specific validation failure(s) per school year. Format: 'SY YYYY: reason1; reason2 | SY YYYY: reason'. Blank if valid."),
            ("", ""),
            ("EXCLUSION REASON SUB-TAGS", ""),
            ("Sub-tag", "Meaning"),
            ("missing grade level(s)", "School did not assess all three grade levels (G1, G2, G3) at that endpoint"),
            ("only N groups", "Fewer than 4 of 6 grade-language groups reported data at that endpoint"),
            ("min N assessed/group", "At least one reporting group had fewer than 15 assessed learners at that endpoint"),
            ("unstable count (>25% change)", "Total assessed count changed by more than 25% between BoSY and EoSY"),
            ("no valid groups", "No grade-language group had any usable data at that endpoint"),
            ("not in dataset / no data", "School was not present in the CRLA data for that endpoint"),
        ]
        for r, (field, value) in enumerate(notes, 1):
            c1 = ws.cell(row=r, column=1, value=field)
            c2 = ws.cell(row=r, column=2, value=value)
            c1.font = Font(bold=True)
            c2.alignment = Alignment(wrap_text=True)
        # Bold section headers
        section_rows = [1, 11, 16, 17, 25, 26, 40, 41]
        for row_idx in section_rows:
            if row_idx <= len(notes):
                for col in (1, 2):
                    ws.cell(row=row_idx, column=col).font = Font(bold=True)
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 100

    print(f"\n✓ Wrote {OUTPUT_PATH}")
    print(f"  Sheet 1: {len(ref)} schools ({n_fc_in_ref} tagged 1st Cycle, {n_valid_composite} valid for composite)")
    print(f"  Sheet 2: Notes")


if __name__ == "__main__":
    main()
